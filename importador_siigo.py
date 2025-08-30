import pandas as pd
import customtkinter as ctk
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import os
import sys
import openpyxl
from openpyxl.styles import numbers
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime
import logging
from typing import Tuple, Optional


class FileManager:
    """Maneja la carga y validación de archivos"""
    
    @staticmethod
    def obtener_ruta_recurso(nombre_archivo: str) -> str:
        """Obtiene la ruta del recurso considerando PyInstaller"""
        if hasattr(sys, '_MEIPASS'):
            return os.path.join(sys._MEIPASS, nombre_archivo)
        return os.path.join(os.path.abspath("."), nombre_archivo)
    
    @staticmethod
    def cargar_hoja_con_columnas(archivo: str, columnas_esperadas: list) -> pd.DataFrame:
        """Carga un archivo Excel buscando las columnas esperadas"""
        try:
            if archivo.lower().endswith(".xls"):
                with open(archivo, "rb") as f:
                    inicio = f.read(1024)
                if b"<table" in inicio.lower():
                    df_list = pd.read_html(archivo)
                    for df in df_list:
                        if all(col in df.columns for col in columnas_esperadas):
                            return df
                    raise ValueError(f"No se encontró una tabla con las columnas requeridas en {archivo}.")
                else:
                    xls = pd.ExcelFile(archivo, engine="xlrd")
            else:
                xls = pd.ExcelFile(archivo, engine="openpyxl")

            for nombre_hoja in xls.sheet_names:
                df = pd.read_excel(xls, sheet_name=nombre_hoja, 
                                 engine='openpyxl' if archivo.endswith('.xlsx') else 'xlrd')
                if all(col in df.columns for col in columnas_esperadas):
                    return df
            
            raise ValueError(f"No se encontró una hoja con las columnas requeridas en {archivo}.")
        except Exception as e:
            logging.error("Error cargando hoja desde %s: %s", archivo, e)
            raise


class DataProcessor:
    """Procesa y transforma los datos de los reportes"""
    
    @staticmethod
    def procesar_reporte1(df: pd.DataFrame) -> pd.DataFrame:
        """Procesa el reporte 1 (Productos)"""
        # Filtrar registros con valor_total != 0
        df = df[df["valor_total"] != 0].copy()
        
        # Calcular valor unitario
        df["Valor unitario"] = df["valor_total"] / df["cantidad"]
        
        # Renombrar columnas
        df = df.rename(columns={
            "factura": "Consecutivo",
            "codigo": "Código producto",
            "referencia": "Descripción producto",
            "cantidad": "Cantidad producto"
        })
        
        # Seleccionar columnas necesarias
        df = df[["Consecutivo", "Código producto", "Descripción producto", "Cantidad producto", "Valor unitario"]]
        df["Consecutivo"] = df["Consecutivo"].astype(str)
        
        return df
    
    @staticmethod
    def procesar_reporte2(df: pd.DataFrame) -> pd.DataFrame:
        """Procesa el reporte 2 (Facturas)"""
        # Renombrar columnas
        df = df.rename(columns={
            "NitEmpresa": "Identificación tercero",
            "f_fact": "Fecha de elaboración",
            "numero": "Consecutivo",
            "total": "Valor Forma de Pago"
        })
        
        # Seleccionar columnas necesarias
        df = df[["Consecutivo", "Identificación tercero", "Fecha de elaboración", "Valor Forma de Pago"]]
        df["Consecutivo"] = df["Consecutivo"].astype(str)
        
        return df
    
    @staticmethod
    def aplicar_filtro_usuario(df: pd.DataFrame, usuario_filtro: str, 
                              filtro_exacto: bool = False, 
                              case_sensitive: bool = False) -> Tuple[pd.DataFrame, str, str]:
        """Aplica filtro de usuario con opciones configurables"""
        if not usuario_filtro or "usuario" not in df.columns:
            return df, "Sin filtro aplicado", "info"
        
        df_original_count = len(df)
        usuario_filtro = usuario_filtro.strip()
        
        try:
            # Limpiar datos de usuario
            df_filtrado = df.copy()
            df_filtrado["usuario"] = df_filtrado["usuario"].fillna("").astype(str)
            
            if filtro_exacto:
                if case_sensitive:
                    mask = df_filtrado["usuario"] == usuario_filtro
                else:
                    mask = df_filtrado["usuario"].str.lower() == usuario_filtro.lower()
            else:
                if case_sensitive:
                    mask = df_filtrado["usuario"].str.contains(usuario_filtro, na=False, regex=False)
                else:
                    mask = df_filtrado["usuario"].str.contains(usuario_filtro, case=False, na=False, regex=False)
            
            df_resultado = df_filtrado[mask]
            df_final_count = len(df_resultado)
            
            # Crear mensaje informativo
            if df_final_count == 0:
                mensaje = f"⚠️ Filtro '{usuario_filtro}' no encontró coincidencias"
                tipo = "warning"
            elif df_final_count == df_original_count:
                mensaje = f"ℹ️ Filtro '{usuario_filtro}' no afectó los datos"
                tipo = "info"
            else:
                mensaje = f"✅ Filtro '{usuario_filtro}': {df_original_count} → {df_final_count} registros"
                tipo = "success"
            
            return df_resultado, mensaje, tipo
            
        except Exception as e:
            return df, f"❌ Error en filtro: {str(e)}", "error"
    
    @staticmethod
    def combinar_reportes(r1: pd.DataFrame, r2: pd.DataFrame) -> pd.DataFrame:
        """Combina los dos reportes"""
        return pd.merge(r1, r2, on="Consecutivo", how="left")
    
    @staticmethod
    def limpiar_datos(df: pd.DataFrame) -> pd.DataFrame:
        """Limpia y prepara los datos finales"""
        # Eliminar registros sin información esencial
        df = df.dropna(subset=["Identificación tercero", "Fecha de elaboración", "Valor Forma de Pago"])
        
        # Filtrar por consecutivos que empiecen con 'E'
        df = df[df["Consecutivo"].astype(str).str.startswith(("E", "e"))]
        
        # Limpiar consecutivos
        df["Consecutivo"] = df["Consecutivo"].astype(str).str.lstrip("Ee")
        
        # Limpiar identificación tercero
        df["Identificación tercero"] = df["Identificación tercero"].astype(str).str.split("-").str[0]
        
        # Convertir fechas
        df["Fecha de elaboración"] = pd.to_datetime(df["Fecha de elaboración"]).dt.date
        
        return df
    
    @staticmethod
    def preparar_estructura_final(df: pd.DataFrame, copiar_fecha_vencimiento: bool = False) -> pd.DataFrame:
        """Prepara la estructura final para SIIGO"""
        columnas_objetivo = [
            "Tipo de comprobante", "Consecutivo", "Identificación tercero", "Sucursal", 
            "Código centro/subcentro de costos", "Fecha de elaboración", "Sigla Moneda", 
            "Tasa de cambio", "Nombre contacto", "Email Contacto", "Orden de compra", 
            "Orden de entrega", "Fecha orden de entrega", "Código producto", 
            "Descripción producto", "Identificación vendedor", "Código de Bodega", 
            "Cantidad producto", "Valor unitario", "Valor Descuento", "Base AIU",
            "Identificación ingreso para terceros", "Código impuesto cargo", 
            "Código impuesto cargo dos", "Código impuesto retención", "Código ReteICA", 
            "Código ReteIVA", "Código forma de pago", "Valor Forma de Pago", 
            "Fecha Vencimiento", "Observaciones"
        ]
        
        # Agregar columnas faltantes
        for col in columnas_objetivo:
            if col not in df.columns:
                df[col] = ""
        
        # Configurar valores por defecto
        df["Tipo de comprobante"] = 1
        df["Identificación vendedor"] = 807001777
        
        if copiar_fecha_vencimiento:
            df["Fecha Vencimiento"] = df["Fecha de elaboración"]
        
        # Seleccionar solo las columnas objetivo
        df = df[columnas_objetivo]
        
        # Consolidar valor forma de pago por consecutivo
        df['Valor Forma de Pago'] = df.groupby('Consecutivo')['Valor Forma de Pago'].transform('first')
        df.loc[df.duplicated('Consecutivo'), 'Valor Forma de Pago'] = ''
        
        return df


class ExcelExporter:
    """Maneja la exportación a Excel"""
    
    @staticmethod
    def generar_archivo(df: pd.DataFrame, plantilla_path: str) -> str:
        """Genera el archivo Excel final"""
        # Crear carpeta de exportados
        carpeta_exportados = os.path.join(os.getcwd(), "Exportados SIIGO")
        os.makedirs(carpeta_exportados, exist_ok=True)
        
        # Generar nombre de archivo
        fecha_hora = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
        archivo_salida = os.path.join(carpeta_exportados, f"SIIGO_Ingresos_{fecha_hora}.xlsx")
        
        # Cargar plantilla
        wb = openpyxl.load_workbook(plantilla_path)
        ws = wb.active
        ws.delete_rows(2, ws.max_row)
        
        # Escribir datos
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=1):
            for c_idx, value in enumerate(row, start=1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                if r_idx == 1:  # Header styling
                    header_cell = ws.cell(row=1, column=c_idx)
                    if hasattr(header_cell, 'fill'):
                        cell.fill = header_cell.fill.copy()
                    if hasattr(header_cell, 'font'):
                        cell.font = header_cell.font.copy()
        
        # Formato de fechas
        columnas_objetivo = list(df.columns)
        if "Fecha de elaboración" in columnas_objetivo:
            fecha_col = columnas_objetivo.index("Fecha de elaboración") + 1
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, 
                                  min_col=fecha_col, max_col=fecha_col):
                for cell in row:
                    if isinstance(cell.value, datetime):
                        cell.number_format = 'YYYY-MM-DD'
        
        wb.save(archivo_salida)
        return archivo_salida


class ModernSiigoApp:
    """Aplicación principal con interfaz moderna"""
    
    def __init__(self):
        self.setup_logging()
        self.setup_variables()
        self.setup_window()
        self.create_widgets()
        
    def setup_logging(self):
        """Configurar logging"""
        os.chdir(os.path.dirname(sys.executable if getattr(sys, 'frozen', False) else __file__))
        log_path = os.path.join(os.getcwd(), "siigo_log.txt")
        logging.basicConfig(
            filename=log_path,
            level=logging.INFO,
            format="%(asctime)s - %(levelname)s - %(message)s"
        )
        
    def setup_variables(self):
        """Configurar variables globales"""
        self.archivo1 = ""
        self.archivo2 = ""
        self.plantilla = FileManager.obtener_ruta_recurso("plantilla_siigo.xlsx")
        
        # Variables de configuración
        self.var_filtro_exacto = None
        self.var_case_sensitive = None
        self.var_fecha_vencimiento = None
        
        # Colores del tema
        self.colors = {
            'primary': '#2E86C1',
            'secondary': '#1B7B3A',
            'success': '#1B7B3A',
            'danger': '#EC7063',
            'warning': '#F7DC6F',
            'light': '#F8F9FA',
            'dark': '#2C3E50',
            'white': '#FFFFFF'
        }
        
    def setup_window(self):
        """Configurar la ventana principal"""
        # Configuración de tema global
        ctk.set_appearance_mode("dark")
        ctk.set_default_color_theme("blue")
        
        self.root = ctk.CTk()
        self.root.title("🚀 Herramienta de Importación SIIGO - v2")
        self.root.geometry("800x600")
        
        # Centrar ventana
        self.center_window()
        
        # Intentar cargar ícono
        try:
            self.root.iconbitmap(FileManager.obtener_ruta_recurso("icono.ico"))
        except:
            pass

    def center_window(self):
        """Centrar la ventana en la pantalla"""
        self.root.update_idletasks()
        width = self.root.winfo_width()
        height = self.root.winfo_height()
        pos_x = (self.root.winfo_screenwidth() // 2) - (width // 2)
        pos_y = (self.root.winfo_screenheight() // 2) - (height // 2)
        self.root.geometry(f'{width}x{height}+{pos_x}+{pos_y}')

    def create_widgets(self):
        """Crear todos los widgets de la interfaz"""
        # Frame principal
        main_frame = ctk.CTkFrame(self.root)
        main_frame.pack(expand=True, fill="both", padx=20, pady=20)
        
        # Header
        self.create_header(main_frame)
        
        # Sección de archivos
        self.create_file_section(main_frame)
        
        # Sección de configuración
        self.create_config_section(main_frame)
        
        # Botón ejecutar
        self.create_execute_section(main_frame)
        
        # Footer
        self.create_footer(main_frame)

    def create_header(self, parent):
        """Crear el header de la aplicación"""
        header_frame = ctk.CTkFrame(parent, fg_color="transparent")
        header_frame.pack(fill="x", pady=(0, 20))
        
        title_label = ctk.CTkLabel(header_frame,
                                  text="Herramienta de Importación SIIGO - v2",
                                  font=ctk.CTkFont(size=24, weight="bold"))
        title_label.pack()
        
        subtitle_label = ctk.CTkLabel(header_frame,
                                     text="Procesa y combina reportes para importar a SIIGO",
                                     font=ctk.CTkFont(size=12))
        subtitle_label.pack(pady=(5, 0))

    def create_file_section(self, parent):
        """Crear la sección de selección de archivos"""
        files_frame = ctk.CTkFrame(parent)
        files_frame.pack(fill="x", pady=(0, 10))
        
        # Título de sección
        section_title = ctk.CTkLabel(files_frame,
                                   text="📁 Selección de Archivos",
                                   font=ctk.CTkFont(size=16, weight="bold"))
        section_title.pack(pady=(10, 15))
        
        # Reporte 1
        r1_frame = ctk.CTkFrame(files_frame, fg_color="transparent")
        r1_frame.pack(fill="x", padx=20, pady=(0, 10))
        
        self.btn_r1 = ctk.CTkButton(r1_frame,
                                   text="📊 Cargar Reporte 1 (Productos)",
                                   corner_radius=10,
                                   width=200,
                                   command=lambda: self.seleccionar_archivo("r1"))
        self.btn_r1.pack(side="left")
        
        self.lbl_r1_status = ctk.CTkLabel(r1_frame,
                                         text="⏳ Esperando archivo...",
                                         font=ctk.CTkFont(size=12))
        self.lbl_r1_status.pack(side="left", padx=(15, 0))
        
        # Reporte 2
        r2_frame = ctk.CTkFrame(files_frame, fg_color="transparent")
        r2_frame.pack(fill="x", padx=20, pady=(0, 15))
        
        self.btn_r2 = ctk.CTkButton(r2_frame,
                                   text="📋 Cargar Reporte 2 (Facturas)",
                                   corner_radius=10,
                                   width=200,
                                   command=lambda: self.seleccionar_archivo("r2"))
        self.btn_r2.pack(side="left")
        
        self.lbl_r2_status = ctk.CTkLabel(r2_frame,
                                         text="⏳ Esperando archivo...",
                                         font=ctk.CTkFont(size=12))
        self.lbl_r2_status.pack(side="left", padx=(15, 0))

    def create_config_section(self, parent):
        """Crear la sección de configuración"""
        config_frame = ctk.CTkFrame(parent)
        config_frame.pack(fill="x", pady=(10, 10))
        
        # Título de sección
        section_title = ctk.CTkLabel(config_frame,
                                   text="⚙️ Configuración",
                                   font=ctk.CTkFont(size=16, weight="bold"))
        section_title.pack(pady=(10, 15))
        
        # Filtro por usuario
        user_frame = ctk.CTkFrame(config_frame, fg_color="transparent")
        user_frame.pack(fill="x", padx=20, pady=(0, 10))
        
        user_label = ctk.CTkLabel(user_frame,
                                 text="👤 Filtrar por usuario:",
                                 font=ctk.CTkFont(size=12, weight="bold"))
        user_label.pack(anchor="w")
        
        entry_frame = ctk.CTkFrame(user_frame, fg_color="transparent")
        entry_frame.pack(fill="x", pady=(5, 0))
        
        self.usuario_entry = ctk.CTkEntry(entry_frame,
                                         placeholder_text="Ingresa el nombre de usuario",
                                         width=200)
        self.usuario_entry.pack(side="left")
        
        btn_usuarios = ctk.CTkButton(entry_frame,
                                   text="Ver Usuarios",
                                   width=100,
                                   command=self.mostrar_usuarios_disponibles)
        btn_usuarios.pack(side="left", padx=(10, 0))
        
        # Opciones de filtro
        options_frame = ctk.CTkFrame(user_frame, fg_color="transparent")
        options_frame.pack(fill="x", pady=(10, 0))
        
        self.var_filtro_exacto = ctk.BooleanVar()
        filtro_exacto_check = ctk.CTkCheckBox(options_frame,
                                             text="Coincidencia exacta",
                                             variable=self.var_filtro_exacto)
        filtro_exacto_check.pack(side="left")
        
        self.var_case_sensitive = ctk.BooleanVar()
        case_sensitive_check = ctk.CTkCheckBox(options_frame,
                                              text="Sensible a mayúsculas",
                                              variable=self.var_case_sensitive)
        case_sensitive_check.pack(side="left", padx=(20, 0))
        
        # Label para feedback del filtro
        self.lbl_filtro_info = ctk.CTkLabel(user_frame,
                                           text="",
                                           font=ctk.CTkFont(size=10))
        
        # Checkbox fecha vencimiento
        self.var_fecha_vencimiento = ctk.BooleanVar()
        fecha_check = ctk.CTkCheckBox(config_frame,
                                     text="📅 Copiar Fecha de elaboración a Fecha Vencimiento",
                                     variable=self.var_fecha_vencimiento)
        fecha_check.pack(padx=20, pady=(10, 15))

    def create_execute_section(self, parent):
        """Crear la sección del botón ejecutar"""
        execute_frame = ctk.CTkFrame(parent, fg_color="transparent")
        execute_frame.pack(fill="x", pady=20)
        
        self.btn_execute = ctk.CTkButton(execute_frame,
                                        text="🚀 EJECUTAR PROCESO",
                                        corner_radius=20,
                                        fg_color="green",
                                        hover_color="darkgreen",
                                        width=300,
                                        height=50,
                                        font=ctk.CTkFont(size=16, weight="bold"),
                                        command=self.ejecutar)
        self.btn_execute.pack(anchor="center")
        
        # Label de estado (inicialmente oculto)
        self.status_label = ctk.CTkLabel(execute_frame,
                                        text="",
                                        font=ctk.CTkFont(size=12))

    def create_footer(self, parent):
        """Crear el footer"""
        footer_frame = ctk.CTkFrame(parent, fg_color="transparent")
        footer_frame.pack(fill="x", pady=(20, 0))
        
        footer_label = ctk.CTkLabel(footer_frame,
                                   text="💡 Herramienta SIIGO v2 | Revisa siigo_log.txt para detalles",
                                   font=ctk.CTkFont(size=10))
        footer_label.pack(side="left")
        
        # Switch de tema
        self.switch_var = ctk.StringVar(value="dark")
        theme_switch = ctk.CTkSwitch(footer_frame,
                                   text="Modo Oscuro",
                                   variable=self.switch_var,
                                   onvalue="dark",
                                   offvalue="light",
                                   command=self.toggle_tema)
        theme_switch.pack(side="right")

    def toggle_tema(self):
        """Cambiar tema de la aplicación"""
        modo = self.switch_var.get()
        ctk.set_appearance_mode(modo)
        logging.info(f"Cambiado a modo: {modo}")

    def seleccionar_archivo(self, tipo: str):
        """Seleccionar archivos con feedback visual"""
        tipos_archivo = [("Excel files", "*.xlsx"), ("Excel files", "*.xls")]
        titulo = f"Seleccionar {'Reporte de Productos' if tipo == 'r1' else 'Reporte de Facturas'}"
        ruta = filedialog.askopenfilename(title=titulo, filetypes=tipos_archivo)
        
        if ruta:
            nombre_archivo = os.path.basename(ruta)
            
            if tipo == "r1":
                self.archivo1 = ruta
                self.lbl_r1_status.configure(text=f"✅ {nombre_archivo}")
                logging.info("Reporte 1 cargado: %s", self.archivo1)
                
            elif tipo == "r2":
                self.archivo2 = ruta
                self.lbl_r2_status.configure(text=f"✅ {nombre_archivo}")
                logging.info("Reporte 2 cargado: %s", self.archivo2)

    def mostrar_usuarios_disponibles(self):
        """Mostrar usuarios disponibles en el Reporte 2"""
        if not self.archivo2:
            messagebox.showwarning("Advertencia", "Primero carga el Reporte 2 (Facturas)")
            return
        
        try:
            df = FileManager.cargar_hoja_con_columnas(self.archivo2, ["usuario"])
            
            if "usuario" not in df.columns:
                messagebox.showinfo("Información", "No se encontró la columna 'usuario' en el Reporte 2")
                return
            
            usuarios = df["usuario"].dropna().unique()
            usuarios = sorted([str(u).strip() for u in usuarios if str(u).strip()])
            
            if not usuarios:
                messagebox.showinfo("Información", "No se encontraron usuarios en el Reporte 2")
                return
            
            # Crear ventana para mostrar usuarios
            self.crear_ventana_usuarios(usuarios)
            
        except Exception as e:
            messagebox.showerror("Error", f"Error al cargar usuarios: {str(e)}")

    def crear_ventana_usuarios(self, usuarios: list):
        """Crear ventana para seleccionar usuarios"""
        ventana = ctk.CTkToplevel(self.root)
        ventana.title("Usuarios Disponibles")
        ventana.geometry("400x500")
        
        # Título
        titulo = ctk.CTkLabel(ventana,
                             text=f"Usuarios encontrados ({len(usuarios)}):",
                             font=ctk.CTkFont(size=14, weight="bold"))
        titulo.pack(pady=(20, 10))
        
        # Scrollable frame para usuarios
        scrollable_frame = ctk.CTkScrollableFrame(ventana, width=350, height=350)
        scrollable_frame.pack(pady=10)
        
        self.usuario_seleccionado = None
        
        def seleccionar_usuario(usuario):
            self.usuario_seleccionado = usuario
            for widget in scrollable_frame.winfo_children():
                if isinstance(widget, ctk.CTkButton):
                    widget.configure(fg_color=("gray75", "gray25"))
            # Resaltar el seleccionado
            for widget in scrollable_frame.winfo_children():
                if isinstance(widget, ctk.CTkButton) and widget.cget("text") == usuario:
                    widget.configure(fg_color="green")
        
        # Crear botones para cada usuario
        for usuario in usuarios:
            btn = ctk.CTkButton(scrollable_frame,
                               text=usuario,
                               command=lambda u=usuario: seleccionar_usuario(u),
                               width=300)
            btn.pack(pady=2)
        
        # Botón confirmar
        def confirmar_seleccion():
            if hasattr(self, 'usuario_seleccionado') and self.usuario_seleccionado:
                self.usuario_entry.delete(0, 'end')
                self.usuario_entry.insert(0, self.usuario_seleccionado)
                ventana.destroy()
        
        btn_confirmar = ctk.CTkButton(ventana,
                                     text="Usar Usuario Seleccionado",
                                     command=confirmar_seleccion,
                                     fg_color="green")
        btn_confirmar.pack(pady=10)

    def show_status(self, message: str):
        """Mostrar mensaje de estado"""
        self.status_label.configure(text=message)
        self.status_label.pack(pady=(10, 0))
        self.root.update()

    def hide_status(self):
        """Ocultar mensaje de estado"""
        self.status_label.pack_forget()
        self.root.update()

    def ejecutar(self):
        """Ejecutar el proceso principal"""
        try:
            # Limpiar info previa
            if hasattr(self, 'lbl_filtro_info'):
                self.lbl_filtro_info.pack_forget()
            
            # Validaciones iniciales
            if not self.archivo1 or not self.archivo2 or not self.plantilla:
                messagebox.showerror("❌ Error", "Debes cargar todos los archivos requeridos.")
                logging.error("Faltan archivos por cargar.")
                return
            
            if not os.path.exists(self.archivo1):
                raise FileNotFoundError(f"El archivo Reporte 1 no fue encontrado: {self.archivo1}")
            if not os.path.exists(self.archivo2):
                raise FileNotFoundError(f"El archivo Reporte 2 no fue encontrado: {self.archivo2}")

            # Procesar datos
            self.show_status("🔄 Iniciando procesamiento...")
            
            # Cargar Reporte 1
            self.show_status("📊 Cargando Reporte 1...")
            columnas_r1 = ["factura", "codigo", "referencia", "cantidad", "valor_total"]
            r1 = FileManager.cargar_hoja_con_columnas(self.archivo1, columnas_r1)
            r1 = DataProcessor.procesar_reporte1(r1)
            logging.info("Reporte 1 procesado con %d registros.", len(r1))

            # Cargar Reporte 2
            self.show_status("📋 Cargando Reporte 2...")
            columnas_r2 = ["NitEmpresa", "f_fact", "numero", "total"]
            r2 = FileManager.cargar_hoja_con_columnas(self.archivo2, columnas_r2)

            # Aplicar filtro de usuario
            usuario_filtro = self.usuario_entry.get().strip()
            if usuario_filtro:
                self.show_status(f"👤 Aplicando filtro de usuario: {usuario_filtro}")
                r2, filtro_mensaje, filtro_tipo = DataProcessor.aplicar_filtro_usuario(
                    r2, usuario_filtro,
                    self.var_filtro_exacto.get(),
                    self.var_case_sensitive.get()
                )
                
                # Mostrar resultado del filtro
                color_mensaje = {
                    "success": "green",
                    "warning": "orange",
                    "error": "red",
                    "info": "blue"
                }.get(filtro_tipo, "gray")
                
                self.lbl_filtro_info.configure(text=filtro_mensaje, text_color=color_mensaje)
                self.lbl_filtro_info.pack(pady=(10, 0))
                
                logging.info("Filtro de usuario aplicado: %s", filtro_mensaje)
                
                # Si no hay registros después del filtro, mostrar advertencia
                if len(r2) == 0:
                    self.hide_status()
                    respuesta = messagebox.askyesno(
                        "Sin Resultados", 
                        f"El filtro de usuario '{usuario_filtro}' eliminó todos los registros.\n\n"
                        "¿Deseas continuar sin filtro de usuario?")
                    if respuesta:
                        # Recargar sin filtro
                        self.show_status("🔄 Recargando sin filtro...")
                        r2 = FileManager.cargar_hoja_con_columnas(self.archivo2, columnas_r2)
                        self.lbl_filtro_info.configure(text="ℹ️ Procesando sin filtro de usuario", 
                                                     text_color="blue")
                    else:
                        return

            # Procesar Reporte 2
            self.show_status("🔧 Procesando Reporte 2...")
            r2 = DataProcessor.procesar_reporte2(r2)
            logging.info("Reporte 2 procesado con %d registros.", len(r2))

            # Combinar reportes
            self.show_status("🔗 Combinando reportes...")
            df = DataProcessor.combinar_reportes(r1, r2)
            logging.info("Registros después del merge: %d", len(df))

            # Verificar registros sin coincidencia
            registros_sin_coincidencia = df[df["Identificación tercero"].isna()]
            if len(registros_sin_coincidencia) > 0:
                logging.warning("Se encontraron %d registros sin coincidencia en R2", 
                              len(registros_sin_coincidencia))

            # Limpiar datos
            self.show_status("🧹 Limpiando datos...")
            df = DataProcessor.limpiar_datos(df)

            if len(df) == 0:
                self.hide_status()
                mensaje_error = "No quedaron registros después de aplicar los filtros.\n\n"
                mensaje_error += "Posibles causas:\n"
                mensaje_error += "• El filtro de usuario es muy restrictivo\n"
                mensaje_error += "• No hay consecutivos que empiecen con 'E'\n"
                mensaje_error += "• No hay coincidencias entre ambos reportes"
                raise ValueError(mensaje_error)

            # Preparar estructura final
            self.show_status("📝 Preparando estructura final...")
            df = DataProcessor.preparar_estructura_final(df, self.var_fecha_vencimiento.get())

            # Generar archivo
            self.show_status("💾 Generando archivo Excel...")
            archivo_salida = ExcelExporter.generar_archivo(df, self.plantilla)
            
            # Ocultar estado y mostrar resultado
            self.hide_status()
            logging.info("Archivo generado correctamente: %s", archivo_salida)
            
            # Mensaje de éxito
            mensaje_exito = f"""¡Proceso completado exitosamente!

📊 Registros procesados: {len(df):,}
📁 Archivo generado: {os.path.basename(archivo_salida)}
📂 Ubicación: {os.path.dirname(archivo_salida)}

El archivo está listo para importar en SIIGO."""
            
            messagebox.showinfo("¡Éxito!", mensaje_exito)

        except Exception as e:
            self.hide_status()
            logging.exception("Error durante la ejecución")
            messagebox.showerror("Error", f"Ocurrió un error durante el procesamiento:\n\n{str(e)}")

    def run(self):
        """Ejecutar la aplicación"""
        self.root.mainloop()


def main():
    """Función principal"""
    try:
        app = ModernSiigoApp()
        app.run()
    except Exception as e:
        logging.exception("Error crítico en la aplicación")
        messagebox.showerror("Error Crítico", f"Error al iniciar la aplicación:\n\n{str(e)}")


if __name__ == "__main__":
    main()

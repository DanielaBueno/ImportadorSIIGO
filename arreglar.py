import pandas as pd
import customtkinter as ctk
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import os
import sys
import openpyxl
from openpyxl.styles import numbers
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime
import logging

# Configuración de logging
os.chdir(os.path.dirname(sys.executable if getattr(sys, 'frozen', False) else __file__))
log_path = os.path.join(os.getcwd(), "siigo_log.txt")
logging.basicConfig(
    filename=log_path,
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s - %(message)s"
)

def obtener_ruta_recurso(nombre_archivo):
    if hasattr(sys, '_MEIPASS'):
        return os.path.join(sys._MEIPASS, nombre_archivo)
    return os.path.join(os.path.abspath("."), nombre_archivo)

# Variables globales
archivo1 = ""
archivo2 = ""
plantilla = obtener_ruta_recurso("plantilla_siigo.xlsx")

# 🔹 Configuración de tema global
ctk.set_appearance_mode("dark")   # "dark", "light" o "system" (automático según Windows)
ctk.set_default_color_theme("blue")  # Puedes usar: "blue", "green", "dark-blue"

class ModernSiigoApp:
    def __init__(self, root):
        self.root = root
        self.setup_window()
        self.setup_style()
        self.create_widgets()
        
    def setup_window(self):
        """Configurar la ventana principal"""
        self.root = ctk.CTk()
        self.root.title("🚀 Herramienta de Importación SIIGO - v2")
        self.root.geometry("600x400")
        self.root.configure(bg="#fff4e5")
        
        # Centrar ventana
        self.center_window()
        
        # Intentar cargar ícono
        try:
            self.root.iconbitmap(obtener_ruta_recurso("icono.ico"))
        except:
            pass

    def center_window(self):
        """Centrar la ventana en la pantalla"""
        self.root.update_idletasks()
        width = self.root.winfo_width()
        height = self.root.winfo_height()
        pos_x = (self.root.winfo_screenwidth() // 2) - (width // 2)
        pos_y = (self.root.winfo_screenheight() // 2) - (height // 2)
        self.root.geometry(f'{width}x{height}+{pos_x}+{pos_y}')

    def setup_style(self):
        """Configurar estilos personalizados"""
        self.style = ttk.Style()
        
        # Configurar tema
        self.style.theme_use('clam')
        
        # Colores personalizados
        self.colors = {
            'primary': '#2E86C1',
            'secondary': '#1B7B3A',
            'success': '#1B7B3A',
            'danger': '#EC7063',
            'warning': '#F7DC6F',
            'light': '#F8F9FA',
            'dark': '#2C3E50',
            'white': '#FFFFFF'
        }
        
       # Centrar y organizar elementos
        self.main_frame = ctk.CTkFrame(self.root)
        self.main_frame.pack(expand=True, fill="both", padx=20, pady=20)

        # Botones redondeados
        self.btn1 = ctk.CTkButton(self.main_frame, 
                                  text="📂 Cargar Reporte 1", 
                                  corner_radius=20, 
                                  width=200, 
                                  command=self.cargar_reporte1)
        self.btn1.pack(pady=10)

        self.btn2 = ctk.CTkButton(self.main_frame, 
                                  text="📂 Cargar Reporte 2", 
                                  corner_radius=20, 
                                  width=200, 
                                  command=self.cargar_reporte2)
        self.btn2.pack(pady=10)

        self.btn_run = ctk.CTkButton(self.main_frame, 
                                     text="▶ Ejecutar Proceso", 
                                     corner_radius=20, 
                                     fg_color="green", 
                                     hover_color="darkgreen", 
                                     width=200, 
                                     command=self.ejecutar)
        self.btn_run.pack(pady=20)


        # Estilos para botones
        self.style.configure('Primary.TButton',
                           background=self.colors['primary'],
                           foreground='white',
                           font=('Arial', 10, 'bold'),
                           relief='flat',
                           padding=(20, 10))
        
        self.style.configure('Success.TButton',
                           background=self.colors['success'],
                           foreground='white',
                           font=('Arial', 12, 'bold'),
                           relief='flat',
                           padding=(30, 15))
        
        self.style.configure('File.TButton',
                           background=self.colors['secondary'],
                           foreground='white',
                           font=('Arial', 10),
                           relief='flat',
                           padding=(15, 8))
        
        # Efectos hover
        self.style.map('Primary.TButton',
                      background=[('active', '#1B4F72')])
        self.style.map('Success.TButton',
                      background=[('active', '#27AE60')])
        self.style.map('File.TButton',
                      background=[('active', '#138D75')])
        

    def cargar_reporte1(self):
        archivo = filedialog.askopenfilename(filetypes=[("Archivos Excel", "*.xls *.xlsx")])
        if archivo:
            logging.info(f"Reporte 1 cargado: {archivo}")

    def cargar_reporte2(self):
        archivo = filedialog.askopenfilename(filetypes=[("Archivos Excel", "*.xls *.xlsx")])
        if archivo:
            logging.info(f"Reporte 2 cargado: {archivo}")

    def ejecutar(self):
        logging.info("🔄 Proceso ejecutado (simulado)")

    def run(self):
        self.root.mainloop()

if __name__ == "__main__":
    app = App()
    app.run()

    def create_widgets(self):
        """Crear todos los widgets de la interfaz"""
        # Frame principal con padding
        main_frame = tk.Frame(self.root, bg='#f0f0f0', padx=40, pady=30)
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        # Header
        self.create_header(main_frame)
        
        # Sección de archivos
        self.create_file_section(main_frame)
        
        # Sección de configuración
        self.create_config_section(main_frame)
        
        # Botón ejecutar
        self.create_execute_section(main_frame)
        
        # Footer
        self.create_footer(main_frame)

    def create_header(self, parent):
        """Crear el header de la aplicación"""
        header_frame = tk.Frame(parent, bg='#f0f0f0')
        header_frame.pack(fill=tk.X, pady=(0, 10))
        
        # Título principal
        title_label = tk.Label(header_frame,
                              text="Herramienta de Importación SIIGO - v2",
                              font=('Arial', 18, 'bold'),
                              fg=self.colors['dark'],
                              bg='#f0f0f0')
        title_label.pack()
        
        # Subtítulo
        subtitle_label = tk.Label(header_frame,
                                 text="Procesa y combina reportes para importar a SIIGO",
                                 font=('Arial', 10),
                                 fg='#666666',
                                 bg='#f0f0f0')
        subtitle_label.pack(pady=(5, 0))

    def create_file_section(self, parent):
        """Crear la sección de selección de archivos"""
        files_frame = tk.LabelFrame(parent,
                                   text=" 📁 Selección de Archivos ",
                                   font=('Arial', 12, 'bold'),
                                   fg=self.colors['dark'],
                                   bg='#f0f0f0',
                                   padx=20,
                                   pady=15)
        files_frame.pack(fill=tk.X, pady=(0, 10))
        
        # Reporte 1
        r1_frame = tk.Frame(files_frame, bg='#f0f0f0')
        r1_frame.pack(fill=tk.X, pady=(0, 15))
        
        r1_btn = ttk.Button(r1_frame,
                           text="📊 Cargar Reporte 1 (Productos)",
                           style='File.TButton',
                           command=lambda: self.seleccionar_archivo("r1"))
        r1_btn.pack(side=tk.LEFT)
        
        self.lbl_r1_status = tk.Label(r1_frame,
                                     text="⏳ Esperando archivo...",
                                     font=('Arial', 9),
                                     fg='#666666',
                                     bg='#f0f0f0')
        self.lbl_r1_status.pack(side=tk.LEFT, padx=(15, 0))
        
        # Label para nombre de archivo R1
        self.lbl_r1_name = tk.Label(r1_frame,
                                   text="",
                                   font=('Arial', 8),
                                   fg='#1B7B3A',
                                   bg='#f0f0f0',
                                   wraplength=400)
        
        # Reporte 2
        r2_frame = tk.Frame(files_frame, bg='#f0f0f0')
        r2_frame.pack(fill=tk.X)
        
        r2_btn = ttk.Button(r2_frame,
                           text="📋 Cargar Reporte 2 (Facturas)",
                           style='File.TButton',
                           command=lambda: self.seleccionar_archivo("r2"))
        r2_btn.pack(side=tk.LEFT)
        
        self.lbl_r2_status = tk.Label(r2_frame,
                                     text="⏳ Esperando archivo...",
                                     font=('Arial', 9),
                                     fg='#666666',
                                     bg='#f0f0f0')
        self.lbl_r2_status.pack(side=tk.LEFT, padx=(15, 0))
        
        # Label para nombre de archivo R2
        self.lbl_r2_name = tk.Label(r2_frame,
                                   text="",
                                   font=('Arial', 8),
                                   fg='#1B7B3A',
                                   bg='#f0f0f0',
                                   wraplength=400)

    def create_config_section(self, parent):
        """Crear la sección de configuración"""
        config_frame = tk.LabelFrame(parent,
                                    text=" ⚙️ Configuración ",
                                    font=('Arial', 12, 'bold'),
                                    fg=self.colors['dark'],
                                    bg='#f0f0f0',
                                    padx=20,
                                    pady=15)
        config_frame.pack(fill=tk.X, pady=(10, 10))
        
        # Filtro por usuario con mejoras
        user_frame = tk.Frame(config_frame, bg='#f0f0f0')
        user_frame.pack(fill=tk.X, pady=(0, 15))
        
        user_label = tk.Label(user_frame,
                             text="👤 Filtrar por usuario:",
                             font=('Arial', 10, 'bold'),
                             fg=self.colors['dark'],
                             bg='#f0f0f0')
        user_label.pack(anchor=tk.W)
        
        # Frame para entry y opciones
        entry_frame = tk.Frame(user_frame, bg='#f0f0f0')
        entry_frame.pack(fill=tk.X, pady=(5, 0))
        
        self.usuario_entry = tk.Entry(entry_frame,
                                     font=('Arial', 10),
                                     relief='solid',
                                     bd=1,
                                     bg=self.colors['white'],
                                     width=30)
        self.usuario_entry.pack(side=tk.LEFT)
        
        # Botón para mostrar usuarios disponibles
        btn_mostrar_usuarios = ttk.Button(entry_frame,
                                         text="Ver Usuarios",
                                         style='Primary.TButton',
                                         command=self.mostrar_usuarios_disponibles)
        btn_mostrar_usuarios.pack(side=tk.LEFT, padx=(10, 0))
        
        # Opciones de filtro
        filter_options_frame = tk.Frame(user_frame, bg='#f0f0f0')
        filter_options_frame.pack(fill=tk.X, pady=(10, 0))
        
        self.var_filtro_exacto = tk.BooleanVar()
        filtro_exacto_check = tk.Checkbutton(filter_options_frame,
                                            text="Coincidencia exacta",
                                            variable=self.var_filtro_exacto,
                                            font=('Arial', 9),
                                            fg=self.colors['dark'],
                                            bg='#f0f0f0',
                                            activebackground='#f0f0f0')
        filtro_exacto_check.pack(side=tk.LEFT)
        
        self.var_case_sensitive = tk.BooleanVar()
        case_sensitive_check = tk.Checkbutton(filter_options_frame,
                                             text="Sensible a mayúsculas",
                                             variable=self.var_case_sensitive,
                                             font=('Arial', 9),
                                             fg=self.colors['dark'],
                                             bg='#f0f0f0',
                                             activebackground='#f0f0f0')
        case_sensitive_check.pack(side=tk.LEFT, padx=(20, 0))
        
        # Label para feedback del filtro
        self.lbl_filtro_info = tk.Label(user_frame,
                                       text="",
                                       font=('Arial', 8),
                                       fg='#666666',
                                       bg='#f0f0f0',
                                       wraplength=500)
        
        # Checkbox fecha vencimiento
        fecha_frame = tk.Frame(config_frame, bg='#f0f0f0')
        fecha_frame.pack(fill=tk.X, pady=(15, 0))
        
        self.var_fecha_vencimiento = tk.BooleanVar()
        fecha_check = tk.Checkbutton(fecha_frame,
                                    text="📅 Copiar Fecha de elaboración a Fecha Vencimiento",
                                    variable=self.var_fecha_vencimiento,
                                    font=('Arial', 10),
                                    fg=self.colors['dark'],
                                    bg='#f0f0f0',
                                    activebackground='#f0f0f0',
                                    activeforeground=self.colors['primary'])
        fecha_check.pack(anchor=tk.W)

    def create_execute_section(self, parent):
        """Crear la sección del botón ejecutar"""
        execute_frame = tk.Frame(parent, bg='#f0f0f0')
        execute_frame.pack(fill=tk.X, pady=20)
        
        execute_btn = ttk.Button(execute_frame,
                               text="🚀 EJECUTAR PROCESO",
                               style='Success.TButton',
                               command=self.ejecutar)
        execute_btn.pack(anchor=tk.CENTER)
        
        # Barra de progreso (inicialmente oculta)
        self.progress = ttk.Progressbar(execute_frame,
                                       mode='indeterminate',
                                       length=400)
        
        self.status_label = tk.Label(execute_frame,
                                   text="",
                                   font=('Arial', 9),
                                   fg=self.colors['primary'],
                                   bg='#f0f0f0')

    def create_footer(self, parent):
        # Footer con switch de tema
        footer_frame = ctk.CTkFrame(self.root, fg_color="transparent")
        footer_frame.pack(side="bottom", pady=10)

        self.footer = ctk.CTkLabel(footer_frame, 
                                   text="💡 Herramienta SIIGO v2 | Revisa siigo_log.txt para detalles",
                                   font=("Arial", 10))
        self.footer.pack(side="left", padx=10)
        footer_label.pack()

        # 🔹 Interruptor de tema (claro/oscuro)
        self.switch_var = ctk.StringVar(value="dark")
        self.switch = ctk.CTkSwitch(footer_frame, 
                                    text="Modo Oscuro", 
                                    variable=self.switch_var, 
                                    onvalue="dark", 
                                    offvalue="light", 
                                    command=self.toggle_tema)
        self.switch.select()  # activado por defecto en oscuro
        self.switch.pack(side="right", padx=10)

    def toggle_tema(self):
        modo = self.switch_var.get()
        ctk.set_appearance_mode(modo)
        logging.info(f"Cambiado a modo: {modo}")

    def seleccionar_archivo(self, tipo):
        """Seleccionar archivos con mejor feedback visual"""
        global archivo1, archivo2
        
        tipos_archivo = [("Excel files", "*.xlsx"), ("Excel files", "*.xls")]
        ruta = filedialog.askopenfilename(
            title=f"Seleccionar {'Reporte de Productos' if tipo == 'r1' else 'Reporte de Facturas'}",
            filetypes=tipos_archivo
        )
        
        if ruta:
            nombre_archivo = os.path.basename(ruta)
            
            if tipo == "r1":
                archivo1 = ruta
                self.lbl_r1_status.config(text="✅ Archivo cargado correctamente",
                                         fg=self.colors['success'])
                self.lbl_r1_name.config(text=f"📄 {nombre_archivo}")
                self.lbl_r1_name.pack(pady=(5, 0))
                logging.info("Reporte 1 cargado: %s", archivo1)
                
            elif tipo == "r2":
                archivo2 = ruta
                self.lbl_r2_status.config(text="✅ Archivo cargado correctamente",
                                         fg=self.colors['success'])
                self.lbl_r2_name.config(text=f"📄 {nombre_archivo}")
                self.lbl_r2_name.pack(pady=(5, 0))
                logging.info("Reporte 2 cargado: %s", archivo2)

    def mostrar_usuarios_disponibles(self):
        """Mostrar usuarios disponibles en el Reporte 2"""
        global archivo2
        
        if not archivo2:
            messagebox.showwarning("Advertencia", "Primero carga el Reporte 2 (Facturas)")
            return
        
        try:
            # Cargar solo para ver usuarios
            def cargar_hoja_con_columnas_simple(archivo):
                if archivo.lower().endswith(".xls"):
                    with open(archivo, "rb") as f:
                        inicio = f.read(1024)
                    if b"<table" in inicio.lower():
                        df_list = pd.read_html(archivo)
                        return df_list[0] if df_list else pd.DataFrame()
                    else:
                        return pd.read_excel(archivo, engine="xlrd")
                else:
                    return pd.read_excel(archivo, engine="openpyxl")
            
            df = cargar_hoja_con_columnas_simple(archivo2)
            
            if "usuario" not in df.columns:
                messagebox.showinfo("Información", "No se encontró la columna 'usuario' en el Reporte 2")
                return
            
            # Obtener usuarios únicos
            usuarios = df["usuario"].dropna().unique()
            usuarios = sorted([str(u).strip() for u in usuarios if str(u).strip()])
            
            if not usuarios:
                messagebox.showinfo("Información", "No se encontraron usuarios en el Reporte 2")
                return
            
            # Crear ventana para mostrar usuarios
            ventana_usuarios = tk.Toplevel(self.root)
            ventana_usuarios.title("Usuarios Disponibles")
            ventana_usuarios.geometry("400x500")
            ventana_usuarios.configure(bg='#f0f0f0')
            
            # Lista de usuarios
            frame_lista = tk.Frame(ventana_usuarios, bg='#f0f0f0', padx=20, pady=20)
            frame_lista.pack(fill=tk.BOTH, expand=True)
            
            tk.Label(frame_lista,
                    text=f"Usuarios encontrados ({len(usuarios)}):",
                    font=('Arial', 12, 'bold'),
                    bg='#f0f0f0').pack(anchor=tk.W, pady=(0, 10))
            
            # Scrollable listbox
            listbox_frame = tk.Frame(frame_lista, bg='#f0f0f0')
            listbox_frame.pack(fill=tk.BOTH, expand=True)
            
            scrollbar = tk.Scrollbar(listbox_frame)
            scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
            
            listbox = tk.Listbox(listbox_frame, 
                                yscrollcommand=scrollbar.set,
                                font=('Arial', 10),
                                bg='white',
                                selectmode=tk.SINGLE)
            listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
            scrollbar.config(command=listbox.yview)
            
            for usuario in usuarios:
                listbox.insert(tk.END, usuario)
            
            # Botón para seleccionar
            def seleccionar_usuario():
                selection = listbox.curselection()
                if selection:
                    usuario_seleccionado = listbox.get(selection[0])
                    self.usuario_entry.delete(0, tk.END)
                    self.usuario_entry.insert(0, usuario_seleccionado)
                    ventana_usuarios.destroy()
            
            btn_seleccionar = ttk.Button(frame_lista,
                                        text="Usar Usuario Seleccionado",
                                        style='Primary.TButton',
                                        command=seleccionar_usuario)
            btn_seleccionar.pack(pady=(10, 0))
            
        except Exception as e:
            messagebox.showerror("Error", f"Error al cargar usuarios: {str(e)}")

    def aplicar_filtro_usuario_mejorado(self, df, usuario_filtro):
        """Aplicar filtro de usuario con opciones mejoradas"""
        if not usuario_filtro or "usuario" not in df.columns:
            return df, "Sin filtro aplicado"
        
        df_original_count = len(df)
        usuario_filtro = usuario_filtro.strip()
        
        try:
            # Limpiar datos de usuario (eliminar NaN y convertir a string)
            df_filtrado = df.copy()
            df_filtrado["usuario"] = df_filtrado["usuario"].fillna("").astype(str)
            
            if self.var_filtro_exacto.get():
                # Filtro exacto
                if self.var_case_sensitive.get():
                    mask = df_filtrado["usuario"] == usuario_filtro
                else:
                    mask = df_filtrado["usuario"].str.lower() == usuario_filtro.lower()
            else:
                # Filtro con contains
                if self.var_case_sensitive.get():
                    mask = df_filtrado["usuario"].str.contains(usuario_filtro, na=False, regex=False)
                else:
                    mask = df_filtrado["usuario"].str.contains(usuario_filtro, case=False, na=False, regex=False)
            
            df_resultado = df_filtrado[mask]
            df_final_count = len(df_resultado)
            
            # Crear mensaje informativo
            if df_final_count == 0:
                mensaje = f"⚠️ Filtro '{usuario_filtro}' no encontró coincidencias"
                tipo = "warning"
            elif df_final_count == df_original_count:
                mensaje = f"ℹ️ Filtro '{usuario_filtro}' no afectó los datos"
                tipo = "info"
            else:
                mensaje = f"✅ Filtro '{usuario_filtro}': {df_original_count} → {df_final_count} registros"
                tipo = "success"
            
            return df_resultado, mensaje, tipo
            
        except Exception as e:
            return df, f"❌ Error en filtro: {str(e)}", "error"

    def show_progress(self, message):
        """Mostrar barra de progreso y mensaje de estado"""
        self.progress.pack(pady=(10, 5))
        self.progress.start(10)
        self.status_label.config(text=message)
        self.status_label.pack(pady=(5, 0))
        self.root.update()

    def hide_progress(self):
        """Ocultar barra de progreso"""
        self.progress.stop()
        self.progress.pack_forget()
        self.status_label.pack_forget()
        self.root.update()

    def ejecutar(self):
        """Ejecutar el proceso principal con filtro de usuario mejorado"""
        try:
            # Limpiar info previa del filtro
            self.lbl_filtro_info.pack_forget()
            
            # Validaciones iniciales
            if not archivo1 or not archivo2 or not plantilla:
                messagebox.showerror("❌ Error", "Debes cargar todos los archivos requeridos.")
                logging.error("Faltan archivos por cargar.")
                return
            
            if not os.path.exists(archivo1):
                raise FileNotFoundError(f"El archivo Reporte 1 no fue encontrado: {archivo1}")
            if not os.path.exists(archivo2):
                raise FileNotFoundError(f"El archivo Reporte 2 no fue encontrado: {archivo2}")

            # Mostrar progreso
            self.show_progress("🔄 Iniciando procesamiento...")

            # Función para cargar hojas con columnas específicas
            def cargar_hoja_con_columnas(archivo, columnas_esperadas):
                try:
                    if archivo.lower().endswith(".xls"):
                        with open(archivo, "rb") as f:
                            inicio = f.read(1024)
                        if b"<table" in inicio.lower():
                            df_list = pd.read_html(archivo)
                            for df in df_list:
                                if all(col in df.columns for col in columnas_esperadas):
                                    return df
                            raise ValueError(f"No se encontró una tabla con las columnas requeridas en {archivo}.")
                        else:
                            xls = pd.ExcelFile(archivo, engine="xlrd")
                    else:
                        xls = pd.ExcelFile(archivo, engine="openpyxl")

                    for nombre_hoja in xls.sheet_names:
                        df = pd.read_excel(xls, sheet_name=nombre_hoja, 
                                         engine='openpyxl' if archivo.endswith('.xlsx') else 'xlrd')
                        if all(col in df.columns for col in columnas_esperadas):
                            return df
                    
                    raise ValueError(f"No se encontró una hoja con las columnas requeridas en {archivo}.")
                except Exception as e:
                    logging.error("Error cargando hoja desde %s: %s", archivo, e)
                    raise

            # Cargar Reporte 1
            self.show_progress("📊 Cargando Reporte 1...")
            columnas_r1 = ["factura", "codigo", "referencia", "cantidad", "valor_total"]
            r1 = cargar_hoja_con_columnas(archivo1, columnas_r1)
            logging.info("Reporte 1 cargado con %d registros.", len(r1))

            # Procesar Reporte 1
            self.show_progress("🔧 Procesando Reporte 1...")
            r1 = r1[r1["valor_total"] != 0]
            r1["Valor unitario"] = r1["valor_total"] / r1["cantidad"]
            r1 = r1.rename(columns={
                "factura": "Consecutivo",
                "codigo": "Código producto",
                "referencia": "Descripción producto",
                "cantidad": "Cantidad producto"
            })
            r1 = r1[["Consecutivo", "Código producto", "Descripción producto", "Cantidad producto", "Valor unitario"]]
            r1["Consecutivo"] = r1["Consecutivo"].astype(str)

            # Cargar Reporte 2
            self.show_progress("📋 Cargando Reporte 2...")
            columnas_r2 = ["NitEmpresa", "f_fact", "numero", "total"]
            r2 = cargar_hoja_con_columnas(archivo2, columnas_r2)
            logging.info("Reporte 2 cargado con %d registros.", len(r2))

            # APLICAR FILTRO DE USUARIO MEJORADO
            usuario_filtro = self.usuario_entry.get().strip()
            filtro_mensaje = ""
            filtro_tipo = "info"
            
            if usuario_filtro:
                self.show_progress(f"👤 Aplicando filtro de usuario: {usuario_filtro}")
                r2, filtro_mensaje, filtro_tipo = self.aplicar_filtro_usuario_mejorado(r2, usuario_filtro)
                
                # Mostrar resultado del filtro
                color_mensaje = {
                    "success": "#1B7B3A",
                    "warning": "#F39C12",
                    "error": "#E74C3C",
                    "info": "#3498DB"
                }.get(filtro_tipo, "#666666")
                
                self.lbl_filtro_info.config(text=filtro_mensaje, fg=color_mensaje)
                self.lbl_filtro_info.pack(pady=(10, 0))
                
                logging.info("Filtro de usuario aplicado: %s", filtro_mensaje)
                
                # Si no hay registros después del filtro, mostrar advertencia
                if len(r2) == 0:
                    self.hide_progress()
                    respuesta = messagebox.askyesno(
                        "Sin Resultados", 
                        f"El filtro de usuario '{usuario_filtro}' eliminó todos los registros.\n\n"
                        "¿Deseas continuar sin filtro de usuario?")
                    if respuesta:
                        # Recargar sin filtro
                        self.show_progress("🔄 Recargando sin filtro...")
                        r2 = cargar_hoja_con_columnas(archivo2, columnas_r2)
                        self.lbl_filtro_info.config(text="ℹ️ Procesando sin filtro de usuario", 
                                                   fg="#3498DB")
                    else:
                        return

            # Procesar Reporte 2
            self.show_progress("🔧 Procesando Reporte 2...")
            r2 = r2.rename(columns={
                "NitEmpresa": "Identificación tercero",
                "f_fact": "Fecha de elaboración",
                "numero": "Consecutivo",
                "total": "Valor Forma de Pago"
            })
            r2 = r2[["Consecutivo", "Identificación tercero", "Fecha de elaboración", "Valor Forma de Pago"]]
            r2["Consecutivo"] = r2["Consecutivo"].astype(str)

            # Combinar datos
            self.show_progress("🔗 Combinando reportes...")
            df = pd.merge(r1, r2, on="Consecutivo", how="left")
            logging.info("Registros después del merge: %d", len(df))

            # Verificar registros sin coincidencia
            registros_sin_coincidencia = df[df["Identificación tercero"].isna()]
            if len(registros_sin_coincidencia) > 0:
                logging.warning("Se encontraron %d registros sin coincidencia en R2", len(registros_sin_coincidencia))

            # Limpiar datos
            self.show_progress("🧹 Limpiando datos...")
            df = df.dropna(subset=["Identificación tercero", "Fecha de elaboración", "Valor Forma de Pago"])
            df = df[df["Consecutivo"].astype(str).str.startswith(("E", "e"))]
            df["Consecutivo"] = df["Consecutivo"].astype(str).str.lstrip("Ee")
            df["Identificación tercero"] = df["Identificación tercero"].astype(str).str.split("-").str[0]
            df["Fecha de elaboración"] = pd.to_datetime(df["Fecha de elaboración"]).dt.date

            if len(df) == 0:
                self.hide_progress()
                mensaje_error = "No quedaron registros después de aplicar los filtros.\n\n"
                mensaje_error += "Posibles causas:\n"
                mensaje_error += "• El filtro de usuario es muy restrictivo\n"
                mensaje_error += "• No hay consecutivos que empiecen con 'E'\n"
                mensaje_error += "• No hay coincidencias entre ambos reportes"
                raise ValueError(mensaje_error)

            # Preparar estructura final
            self.show_progress("📝 Preparando estructura final...")
            columnas_objetivo = [
                "Tipo de comprobante", "Consecutivo", "Identificación tercero", "Sucursal", 
                "Código centro/subcentro de costos", "Fecha de elaboración", "Sigla Moneda", 
                "Tasa de cambio", "Nombre contacto", "Email Contacto", "Orden de compra", 
                "Orden de entrega", "Fecha orden de entrega", "Código producto", 
                "Descripción producto", "Identificación vendedor", "Código de Bodega", 
                "Cantidad producto", "Valor unitario", "Valor Descuento", "Base AIU",
                "Identificación ingreso para terceros", "Código impuesto cargo", 
                "Código impuesto cargo dos", "Código impuesto retención", "Código ReteICA", 
                "Código ReteIVA", "Código forma de pago", "Valor Forma de Pago", 
                "Fecha Vencimiento", "Observaciones"
            ]

            for col in columnas_objetivo:
                if col not in df.columns:
                    df[col] = ""

            df["Tipo de comprobante"] = 1
            df["Identificación vendedor"] = 807001777

            if self.var_fecha_vencimiento.get():
                df["Fecha Vencimiento"] = df["Fecha de elaboración"]
                logging.info("Fecha de elaboración copiada a Fecha Vencimiento.")

            df = df[columnas_objetivo]
            df['Valor Forma de Pago'] = df.groupby('Consecutivo')['Valor Forma de Pago'].transform('first')
            df.loc[df.duplicated('Consecutivo'), 'Valor Forma de Pago'] = ''

            # Generar archivo
            self.show_progress("💾 Generando archivo Excel...")
            carpeta_exportados = os.path.join(os.getcwd(), "Exportados SIIGO")
            os.makedirs(carpeta_exportados, exist_ok=True)

            fecha_hora = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
            archivo_salida = os.path.join(carpeta_exportados, f"SIIGO_Ingresos_{fecha_hora}.xlsx")

            wb = openpyxl.load_workbook(plantilla)
            ws = wb.active
            ws.delete_rows(2, ws.max_row)

            for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=1):
                for c_idx, value in enumerate(row, start=1):
                    cell = ws.cell(row=r_idx, column=c_idx, value=value)
                    if r_idx == 1:  # Header styling
                        header_cell = ws.cell(row=1, column=c_idx)
                        if hasattr(header_cell, 'fill'):
                            cell.fill = header_cell.fill.copy()
                        if hasattr(header_cell, 'font'):
                            cell.font = header_cell.font.copy()

            # Formato de fechas
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, 
                                  min_col=columnas_objetivo.index("Fecha de elaboración") + 1, 
                                  max_col=columnas_objetivo.index("Fecha de elaboración") + 1):
                for cell in row:
                    if isinstance(cell.value, datetime):
                        cell.number_format = 'YYYY-MM-DD'

            wb.save(archivo_salida)
            
            # Ocultar progreso y mostrar resultado
            self.hide_progress()
            logging.info("Archivo generado correctamente: %s", archivo_salida)
            
            # Mensaje de éxito con información del filtro
            mensaje_exito = f"""¡Proceso completado exitosamente!

📊 Registros procesados: {len(df):,}
📁 Archivo generado: {os.path.basename(archivo_salida)}
📂 Ubicación: {carpeta_exportados}

{filtro_mensaje if filtro_mensaje else ''}

El archivo está listo para importar en SIIGO."""
            
            messagebox.showinfo("¡Éxito!", mensaje_exito)

        except Exception as e:
            self.hide_progress()
            logging.exception("Error durante la ejecución")
            messagebox.showerror("Error", f"Ocurrió un error durante el procesamiento:\n\n{str(e)}")

# Crear y ejecutar la aplicación
if __name__ == "__main__":
    root = tk.Tk()
    app = ModernSiigoApp(root)
    root.mainloop()
